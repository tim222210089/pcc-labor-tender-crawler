from __future__ import annotations

import argparse
import html
import re
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable
from urllib.parse import parse_qsl, urlencode, urljoin, urlsplit, urlunsplit

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


BASE_URL = "https://web.pcc.gov.tw"
SEARCH_PATH = "/prkms/tender/common/advanced/readTenderAdvanced"
AWARD_SEARCH_PATH = "/prkms/tender/common/agent/readTenderAgent"
TODAY_PARAMS = {
    "tenderType": "TENDER_DECLARATION",
    "tenderWay": "TENDER_WAY_ALL_DECLARATION",
    "dateType": "isNow",
    "firstSearch": "true",
    "searchType": "advanced",
    "pageSize": "100",
}
DATE_POST_BASE = {
    "tenderType": "TENDER_DECLARATION",
    "tenderWay": "TENDER_WAY_ALL_DECLARATION",
    "dateType": "isDate",
    "searchType": "advanced",
    "pageSize": "100",
    "firstSearch": "false",
    "isBinding": "N",
    "isLogIn": "N",
}
AWARD_POST_BASE = {
    "tenderWay": "TENDER_WAY_ALL_DECLARATION",
    "radProctrgCate": "RAD_PROCTRG_CATE_3",
    "pageSize": "100",
    "firstSearch": "false",
    "isQuery": "Y",
    "isBinding": "N",
    "isLogIn": "N",
}
TENDER_HEADERS = ["機關名稱", "標案名稱", "公告日期", "截止投標", "預算金額"]
AWARD_HEADERS = ["機關名稱", "標案名稱", "公告日期", "決標金額", "決標公告", "無法決標"]
DEFAULT_KEYWORDS = ["委託", "設計", "監造", "技術服務"]
SUPPORT_ONLY_KEYWORD = "委託"
TITLE_RE = re.compile(r'pageCode2Img\("(?P<title>.*?)"\)')
PAGE_LINK_RE = re.compile(r'href="(?P<href>\?[^"]*d-\d+-p=\d+[^"]*)"')
PAGE_PARAM_RE = re.compile(r"[?&](?P<name>d-\d+-p)=(?P<page>\d+)")


@dataclass(frozen=True)
class TenderRecord:
    agency: str
    title: str
    notice_date: str
    deadline: str
    budget: str
    procurement_type: str

    def as_row(self) -> list[str]:
        return [self.agency, self.title, self.notice_date, self.deadline, self.budget]


@dataclass(frozen=True)
class AwardRecord:
    agency: str
    title: str
    notice_date: str
    award_amount: str
    award_notice: str
    failed_award: str

    def as_row(self) -> list[str]:
        return [
            self.agency,
            self.title,
            self.notice_date,
            self.award_amount,
            self.award_notice,
            self.failed_award,
        ]


def normalize_space(value: str) -> str:
    return re.sub(r"\s+", " ", value.replace("\xa0", " ")).strip()


def roc_date_string(target: date) -> str:
    return f"{target.year - 1911:03d}/{target.month:02d}/{target.day:02d}"


def slash_date_string(target: date) -> str:
    return target.strftime("%Y/%m/%d")


def parse_cli_date(raw: str) -> tuple[str, date | None]:
    if raw.lower() == "today":
        return "today", None
    try:
        parsed = datetime.strptime(raw, "%Y-%m-%d").date()
    except ValueError as exc:
        raise SystemExit(f"--date must be 'today' or YYYY-MM-DD, got: {raw}") from exc
    return raw, parsed


def resolve_target_date(explicit_date: date | None) -> date:
    return explicit_date or date.today()


def fetch_with_retries(
    session: requests.Session,
    method: str,
    url: str,
    *,
    params: dict[str, str] | None = None,
    data: dict[str, str] | None = None,
    retries: int = 4,
    timeout: int = 30,
) -> str:
    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            response = session.request(method, url, params=params, data=data, timeout=timeout)
            response.raise_for_status()
            response.encoding = response.apparent_encoding or response.encoding
            return response.text
        except requests.RequestException as exc:
            last_error = exc
            if attempt == retries:
                break
            time.sleep(1.5 * attempt)
    raise RuntimeError(f"Failed to fetch PCC page after {retries} attempts: {last_error}") from last_error


def extract_title(cell) -> str:
    cell_html = str(cell)
    match = TITLE_RE.search(cell_html)
    if match:
        return normalize_space(html.unescape(match.group("title")))

    text = normalize_space(cell.get_text(" ", strip=True))
    parts = [part for part in text.split(" ") if part]
    if len(parts) >= 2 and re.fullmatch(r"\d{6,}", parts[0]):
        return normalize_space(" ".join(parts[1:]))
    return text


def parse_tender_rows(page_html: str) -> list[TenderRecord]:
    soup = BeautifulSoup(page_html, "html.parser")
    table = soup.find("table", id="atm") or soup.find("table", id="tpam")
    if table is None:
        return []

    records: list[TenderRecord] = []
    for row in table.select("tr.tb_b2, tr.tb_b3"):
        cells = row.find_all("td")
        if len(cells) < 9:
            continue
        procurement_type = normalize_space(cells[5].get_text(" ", strip=True))
        records.append(
            TenderRecord(
                agency=normalize_space(cells[1].get_text(" ", strip=True)),
                title=extract_title(cells[2]),
                notice_date=normalize_space(cells[6].get_text(" ", strip=True)),
                deadline=normalize_space(cells[7].get_text(" ", strip=True)),
                budget=normalize_space(cells[8].get_text(" ", strip=True)),
                procurement_type=procurement_type,
            )
        )
    return records


def parse_award_rows(page_html: str) -> list[AwardRecord]:
    soup = BeautifulSoup(page_html, "html.parser")
    table = soup.find("table", id="atm") or soup.find("table", id="tpam")
    if table is None:
        return []

    records: list[AwardRecord] = []
    for row in table.select("tr.tb_b2, tr.tb_b3"):
        cells = row.find_all("td")
        if len(cells) < 9:
            continue
        records.append(
            AwardRecord(
                agency=normalize_space(cells[1].get_text(" ", strip=True)),
                title=extract_title(cells[2]),
                notice_date=normalize_space(cells[5].get_text(" ", strip=True)),
                award_amount=normalize_space(cells[6].get_text(" ", strip=True)),
                award_notice=normalize_space(cells[7].get_text(" ", strip=True)),
                failed_award=normalize_space(cells[8].get_text(" ", strip=True)),
            )
        )
    return records


def parse_next_page_links(page_html: str) -> list[str]:
    return [html.unescape(match.group("href")) for match in PAGE_LINK_RE.finditer(page_html)]


def build_all_page_links(page_links: list[str]) -> list[str]:
    pages: list[tuple[int, str, str]] = []
    for href in page_links:
        match = PAGE_PARAM_RE.search(href)
        if not match:
            continue
        pages.append((int(match.group("page")), match.group("name"), href))

    if not pages:
        return []

    _, param_name, template = max(pages, key=lambda item: item[0])
    last_page = max(page for page, _, _ in pages)
    return [replace_query_param(template, param_name, str(page)) for page in range(2, last_page + 1)]


def replace_query_param(href: str, name: str, value: str) -> str:
    parsed = urlsplit(href)
    query = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query[name] = value
    return urlunsplit(parsed._replace(query=urlencode(query)))


def unique_tenders(records: Iterable[TenderRecord]) -> list[TenderRecord]:
    seen: set[tuple[str, str, str, str, str]] = set()
    result: list[TenderRecord] = []
    for record in records:
        key = (record.agency, record.title, record.notice_date, record.deadline, record.budget)
        if key in seen:
            continue
        seen.add(key)
        result.append(record)
    return result


def unique_awards(records: Iterable[AwardRecord]) -> list[AwardRecord]:
    seen: set[tuple[str, str, str, str, str, str]] = set()
    result: list[AwardRecord] = []
    for record in records:
        key = (
            record.agency,
            record.title,
            record.notice_date,
            record.award_amount,
            record.award_notice,
            record.failed_award,
        )
        if key in seen:
            continue
        seen.add(key)
        result.append(record)
    return result


def build_session(referer_path: str) -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            "Referer": urljoin(BASE_URL, referer_path),
        }
    )
    return session


def fetch_all_records(target_date: date | None) -> list[TenderRecord]:
    session = build_session(SEARCH_PATH)

    if target_date is None:
        first_html = fetch_with_retries(
            session,
            "GET",
            urljoin(BASE_URL, SEARCH_PATH),
            params=TODAY_PARAMS,
        )
    else:
        payload = dict(DATE_POST_BASE)
        roc = roc_date_string(target_date)
        payload["tenderStartDate"] = roc
        payload["tenderEndDate"] = roc
        first_html = fetch_with_retries(
            session,
            "POST",
            urljoin(BASE_URL, SEARCH_PATH),
            data=payload,
        )

    all_records = parse_tender_rows(first_html)
    visited = {urljoin(BASE_URL, SEARCH_PATH)}
    for href in build_all_page_links(parse_next_page_links(first_html)):
        full_url = urljoin(urljoin(BASE_URL, SEARCH_PATH), href)
        if full_url in visited:
            continue
        visited.add(full_url)
        page_html = fetch_with_retries(session, "GET", full_url)
        all_records.extend(parse_tender_rows(page_html))

    return unique_tenders(all_records)


def fetch_award_records(target_date: date, tender_status: str) -> list[AwardRecord]:
    session = build_session(AWARD_SEARCH_PATH)
    payload = dict(AWARD_POST_BASE)
    slash_date = slash_date_string(target_date)
    payload["tenderStatus"] = tender_status
    payload["awardAnnounceStartDate"] = slash_date
    payload["awardAnnounceEndDate"] = slash_date

    first_html = fetch_with_retries(
        session,
        "POST",
        urljoin(BASE_URL, AWARD_SEARCH_PATH),
        data=payload,
    )

    all_records = parse_award_rows(first_html)
    visited = {urljoin(BASE_URL, AWARD_SEARCH_PATH)}
    for href in build_all_page_links(parse_next_page_links(first_html)):
        full_url = urljoin(urljoin(BASE_URL, AWARD_SEARCH_PATH), href)
        if full_url in visited:
            continue
        visited.add(full_url)
        page_html = fetch_with_retries(session, "GET", full_url)
        all_records.extend(parse_award_rows(page_html))

    return unique_awards(all_records)


def filter_labor_records(records: Iterable[TenderRecord]) -> list[TenderRecord]:
    return [record for record in records if record.procurement_type == "勞務類"]


def filter_keyword_matches(records: Iterable[TenderRecord], keywords: list[str]) -> list[TenderRecord]:
    cleaned = [normalize_space(keyword) for keyword in keywords if normalize_space(keyword)]
    standalone_keywords = [keyword for keyword in cleaned if keyword != SUPPORT_ONLY_KEYWORD]
    support_only_enabled = SUPPORT_ONLY_KEYWORD in cleaned
    result: list[TenderRecord] = []
    for record in records:
        title = normalize_space(record.title)
        if any(keyword in title for keyword in standalone_keywords):
            result.append(record)
            continue
        if support_only_enabled and SUPPORT_ONLY_KEYWORD in title and standalone_keywords:
            continue
        if support_only_enabled and SUPPORT_ONLY_KEYWORD in title and not standalone_keywords:
            result.append(record)
    return result


def configure_sheet(sheet, headers: list[str], rows: list[list[str]]) -> None:
    header_font = Font(name="Arial", bold=True)
    body_font = Font(name="Arial")
    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = fill

    for row in rows:
        sheet.append(row)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font

    widths = [22, 56, 14, 14, 18, 14]
    for idx, width in enumerate(widths[: len(headers)], start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width
    sheet.freeze_panes = "A2"


def write_workbook(
    output_path: Path,
    keyword_rows: list[TenderRecord],
    keyword_awarded_rows: list[AwardRecord],
    keyword_failed_rows: list[AwardRecord],
) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    keyword_sheet = wb.create_sheet("keyword_matches")
    keyword_awarded_sheet = wb.create_sheet("keyword_awarded_tenders")
    keyword_failed_sheet = wb.create_sheet("keyword_failed_awards")

    configure_sheet(keyword_sheet, TENDER_HEADERS, [record.as_row() for record in keyword_rows])
    configure_sheet(keyword_awarded_sheet, AWARD_HEADERS, [record.as_row() for record in keyword_awarded_rows])
    configure_sheet(keyword_failed_sheet, AWARD_HEADERS, [record.as_row() for record in keyword_failed_rows])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_output_path(raw_date: str, explicit_date: date | None, output: str | None) -> Path:
    if output:
        return Path(output).expanduser().resolve()
    suffix = explicit_date.isoformat() if explicit_date else raw_date
    return Path.cwd() / f"pcc_labor_tenders_{suffix}.xlsx"


def parse_keywords(raw: str) -> list[str]:
    parts = [normalize_space(item) for item in raw.split(",")]
    return [item for item in parts if item]


def main() -> int:
    parser = argparse.ArgumentParser(description="Export PCC labor tenders to Excel.")
    parser.add_argument("--date", default="today", help="today or YYYY-MM-DD")
    parser.add_argument("--output", help="output xlsx path")
    parser.add_argument(
        "--keywords",
        default=",".join(DEFAULT_KEYWORDS),
        help="comma-separated title keywords; '委託' does not match by itself unless it is the only keyword provided",
    )
    args = parser.parse_args()

    raw_date, explicit_date = parse_cli_date(args.date)
    output_path = build_output_path(raw_date, explicit_date, args.output)
    keywords = parse_keywords(args.keywords)
    target_date = resolve_target_date(explicit_date)

    records = fetch_all_records(explicit_date)
    labor_records = filter_labor_records(records)
    keyword_records = filter_keyword_matches(labor_records, keywords)
    awarded_rows = fetch_award_records(target_date, "TENDER_STATUS_1")
    failed_rows = fetch_award_records(target_date, "TENDER_STATUS_2")
    keyword_awarded_rows = filter_keyword_matches(awarded_rows, keywords)
    keyword_failed_rows = filter_keyword_matches(failed_rows, keywords)
    write_workbook(output_path, keyword_records, keyword_awarded_rows, keyword_failed_rows)

    print(f"output={output_path}")
    print(f"keyword_matches={len(keyword_records)}")
    print(f"keyword_awarded_tenders={len(keyword_awarded_rows)}")
    print(f"keyword_failed_awards={len(keyword_failed_rows)}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except RuntimeError as exc:
        print(str(exc), file=sys.stderr)
        raise SystemExit(1)
